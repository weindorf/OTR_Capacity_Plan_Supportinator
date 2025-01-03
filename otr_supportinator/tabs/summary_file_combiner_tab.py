from PyQt6.QtWidgets import (QVBoxLayout, QHBoxLayout, QLabel, QPushButton, 
                             QListWidget, QFileDialog, QScrollArea, QWidget,
                             QTableWidget, QTableWidgetItem, QHeaderView, QDialog, 
                             QRadioButton, QDialogButtonBox, QComboBox, QSpinBox,
                             QCheckBox, QGroupBox, QSizePolicy, QMessageBox, 
                             QProgressDialog)
from PyQt6.QtCore import Qt, pyqtSignal, QSize, QRect, QThread, QEventLoop
from PyQt6.QtGui import QColor, QResizeEvent, QDropEvent, QDragEnterEvent, QFontMetrics, QPainter
from .base_tab import BaseTab
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font
import os
import re
import numpy as np
import pandas as pd
import psutil

def find_optimal_chunk_size(sample_data, target_memory_usage=0.1):
    available_memory = psutil.virtual_memory().available
    sample_memory = sample_data.memory_usage(deep=True).sum()
    memory_per_row = sample_memory / len(sample_data)
    target_memory = available_memory * target_memory_usage
    chunk_size = int(target_memory / memory_per_row)
    return max(chunk_size, 1)

class FileCombinerWorker(QThread):
    progress_updated = pyqtSignal(int, str)
    error_occurred = pyqtSignal(str)
    process_completed = pyqtSignal(list, str)
    save_location_requested = pyqtSignal()
    save_location_set = pyqtSignal()

    def __init__(self, file_paths, combinations, planning_week, parent=None):
        super().__init__(parent)
        self.file_paths = file_paths
        self.combinations = combinations
        self.planning_week = planning_week
        self.save_directory = None
        self.combination_row_counts = {}
        self.header_format = None
        self.master_data = None

    def run(self):
        try:
            self.progress_updated.emit(0, "Requesting save location...")
            if not self.get_save_location():
                raise Exception("Save location selection cancelled")

            self.progress_updated.emit(5, "Extracting header format...")
            self.extract_header_format()

            self.progress_updated.emit(10, "Reading input files...")
            self.read_and_process_input_files()

            self.progress_updated.emit(50, "Processing combinations...")
            self.process_combinations()

            self.progress_updated.emit(100, "Process completed.")
            self.process_completed.emit(self.get_combination_names(), self.save_directory)
        except Exception as e:
            self.error_occurred.emit(str(e))

    def extract_header_format(self):
        wb = load_workbook(self.file_paths[0], read_only=True)
        ws = wb.active
        self.header_format = [cell.font for cell in next(ws.rows)]

    def read_and_process_input_files(self):
        self.master_data = pd.DataFrame()
        total_files = len(self.file_paths)
        
        for i, file_path in enumerate(self.file_paths, 1):
            df = pd.read_excel(file_path)
            df['planning_horizon'] = (df['amazon_week'] - self.planning_week) % 52
            self.master_data = pd.concat([self.master_data, df], ignore_index=True)
            self.progress_updated.emit(10 + int(40 * i / total_files), f"Reading input file {i}/{total_files}...")

    def process_combinations(self):
        total_combinations = len(self.combinations)
        for i, combination in enumerate(self.combinations, 1):
            self.progress_updated.emit(50 + int(45 * i / total_combinations), f"Processing combination {i}/{total_combinations}...")
            
            output_file = os.path.join(self.save_directory, f"{combination['title']}.xlsx")
            weeks_for_combination = set(range(combination['start_week'], combination['end_week'] + 1))
            
            filtered_df = self.master_data[self.master_data['planning_horizon'].isin(weeks_for_combination)]
            filtered_df = filtered_df.drop(columns=['planning_horizon'])
            
            sample_data = filtered_df.head(1000) if len(filtered_df) > 1000 else filtered_df
            chunk_size = find_optimal_chunk_size(sample_data)
            
            wb = Workbook(write_only=True)
            ws = wb.create_sheet()

            header = []
            for idx, column_name in enumerate(filtered_df.columns):
                cell = WriteOnlyCell(ws, value=column_name)
                if idx < len(self.header_format):
                    cell.font = Font(
                        name=self.header_format[idx].name,
                        size=self.header_format[idx].size,
                        bold=self.header_format[idx].bold,
                        italic=self.header_format[idx].italic,
                    )
                header.append(cell)

            ws.append(header)

            data_array = filtered_df.values

            for start_row in range(0, len(data_array), chunk_size):
                end_row = min(start_row + chunk_size, len(data_array))
                chunk = data_array[start_row:end_row]
                
                for row in chunk:
                    ws.append(row.tolist())

            wb.save(output_file)
            
            self.combination_row_counts[combination['title']] = len(filtered_df)

        self.progress_updated.emit(95, "Finalizing process...")

    def get_save_location(self):
        self.save_location_requested.emit()
        
        loop = QEventLoop()
        self.save_location_set.connect(loop.quit)
        loop.exec()
        
        return self.save_directory is not None

    def set_save_directory(self, directory):
        if directory:
            self.save_directory = directory
            self.save_location_set.emit()

    def get_combination_names(self):
        return [f"{title}.xlsx with {self.combination_row_counts[title]} rows" for title in self.combination_row_counts]

class FileListWidget(QListWidget):
    files_changed = pyqtSignal()
    file_added = pyqtSignal(str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setDragDropMode(QListWidget.DragDropMode.InternalMove)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def dropEvent(self, event: QDropEvent):
        if event.mimeData().hasUrls():
            event.setDropAction(Qt.DropAction.CopyAction)
            for url in event.mimeData().urls():
                file_path = url.toLocalFile()
                if file_path.lower().endswith(('.xlsx', '.xls')):
                    if not self.file_exists(file_path):
                        self.addItem(file_path)
                        self.file_added.emit(file_path)
            event.accept()
            self.files_changed.emit()
        else:
            super().dropEvent(event)

    def file_exists(self, file_path):
        file_name = os.path.basename(file_path)
        for index in range(self.count()):
            if file_name == os.path.basename(self.item(index).text()):
                return True
        return False

class WrappingLabel(QLabel):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setWordWrap(True)
        self.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Minimum)

    def paintEvent(self, event):
        painter = QPainter(self)
        metrics = QFontMetrics(self.font())
        leading = metrics.leading()
        height = 0
        line_count = 0
        available_width = self.width()

        for line in self.text().split('\n'):
            height += leading
            if not line:
                height += metrics.height()
                continue

            line_width = metrics.horizontalAdvance(line)
            if line_width <= available_width:
                line_rect = QRect(0, height, line_width, metrics.height())
                painter.drawText(line_rect, self.alignment(), line)
                height += metrics.height()
            else:
                # Wrap the line character by character
                current_line = ""
                for char in line:
                    new_line = current_line + char
                    new_width = metrics.horizontalAdvance(new_line)
                    if new_width <= available_width:
                        current_line = new_line
                    else:
                        line_rect = QRect(0, height, metrics.horizontalAdvance(current_line), metrics.height())
                        painter.drawText(line_rect, self.alignment(), current_line)
                        height += metrics.height()
                        current_line = char
                    line_count += 1
                if current_line:
                    line_rect = QRect(0, height, metrics.horizontalAdvance(current_line), metrics.height())
                    painter.drawText(line_rect, self.alignment(), current_line)
                    height += metrics.height()

        self.setMinimumHeight(height)

class SummaryFileCombinerTab(BaseTab):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.file_data = {}
        self.planning_week = None
        self.combinations = []
        self.init_ui()
        
        # Connect the window's resize event to our custom handler
        if self.window():
            self.window().installEventFilter(self)

    def init_ui(self):
        # Create a scroll area
        self.scroll_area = QScrollArea(self)
        self.scroll_area.setWidgetResizable(True)
        self.layout.addWidget(self.scroll_area)

        # Create a widget to hold all the content
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.scroll_area.setWidget(self.content_widget)

        # File drop area
        self.file_list = FileListWidget()
        font_metrics = self.file_list.fontMetrics()
        row_height = font_metrics.height()
        self.file_list.setMinimumHeight(row_height * 6)
        self.file_list.files_changed.connect(self.update_ui_state)
        self.file_list.file_added.connect(self.process_file)
        self.file_list.itemSelectionChanged.connect(self.update_remove_button)
        self.content_layout.addWidget(self.file_list)

        # File buttons layout
        file_buttons_layout = QHBoxLayout()
        self.browse_button = QPushButton("Browse Files")
        self.browse_button.clicked.connect(self.browse_files)
        self.remove_button = QPushButton("Remove Selected")
        self.remove_button.clicked.connect(self.remove_selected_files)
        self.remove_button.setEnabled(False)  # Initially disabled
        self.clear_all_button = QPushButton("Clear All")
        self.clear_all_button.clicked.connect(self.clear_all_files)
        file_buttons_layout.addWidget(self.browse_button)
        file_buttons_layout.addWidget(self.remove_button)
        file_buttons_layout.addWidget(self.clear_all_button)
        self.content_layout.addLayout(file_buttons_layout)

        # Planning week layout
        self.planning_week_widget = QWidget()
        planning_week_layout = QHBoxLayout(self.planning_week_widget)
        planning_week_layout.setContentsMargins(0, 0, 0, 0)
        self.planning_week_label = QLabel("Planning Week: Not Set")
        planning_week_layout.addWidget(self.planning_week_label)
        self.content_layout.addWidget(self.planning_week_widget)

        # Planned Weeks Available table
        self.planned_weeks_table = QTableWidget(self)
        self.setup_table()
        self.content_layout.addWidget(self.planned_weeks_table)
        
        # Combinations area
        self.combinations_widget = QWidget()
        self.combinations_widget.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Preferred)
        combinations_layout = QHBoxLayout(self.combinations_widget)
        combinations_layout.setSpacing(0)
        combinations_layout.setContentsMargins(0, 0, 0, 0)
        default_presets = ["All", "A", "B", "C", "D"]
        for i in range(5):
            combination = self.create_combination_widget(i, default_presets[i])
            combinations_layout.addWidget(combination)
            self.combinations.append(combination)
        self.content_layout.addWidget(self.combinations_widget)

        # Add some vertical spacing
        self.content_layout.addSpacing(20)

        # Generate Combined Files button
        self.generate_button = QPushButton("Generate Combined Files")
        self.generate_button.clicked.connect(self.start_combination_process)
        self.content_layout.addWidget(self.generate_button)

        # Add stretch to push everything to the top
        self.content_layout.addStretch(1)

        # Set initial sizes
        self.initial_width = 900
        self.adjust_layout_width(self.initial_width)

        self.update_ui_state()

    def showEvent(self, event):
        super().showEvent(event)
        # Adjust layout width when the tab is first shown
        if self.window():
            self.adjust_layout_width(self.window().width())

    def eventFilter(self, obj, event):
        if obj == self.window() and event.type() == QResizeEvent.Type.Resize:
            self.adjust_layout_width(obj.contentsRect().width())
        return super().eventFilter(obj, event)

    def adjust_layout_width(self, window_width):
        content_width = window_width - 60
        if hasattr(self, 'scroll_area') and self.scroll_area.verticalScrollBar().isVisible():
            content_width -= self.scroll_area.verticalScrollBar().width()
        
        self.content_widget.setFixedWidth(content_width)
        
        if hasattr(self, 'file_list'):
            self.file_list.setFixedWidth(content_width - 20)
        
        if hasattr(self, 'planned_weeks_table'):
            self.planned_weeks_table.setFixedWidth(content_width - 20)
        
        self.update_combination_widths(content_width)

    def update_combination_widths(self, total_width=None):
        if not hasattr(self, 'combinations_widget') or self.combinations_widget is None:
            return
        
        if total_width is None:
            total_width = self.combinations_widget.width()
        
        combination_width = (total_width - 40) // 5  # 40 for some padding
        for combination in self.combinations:
            combination.setFixedWidth(combination_width)

    def clear_all_files(self):
        self.file_list.clear()
        self.file_data.clear()
        self.planning_week = None
        self.planning_week_label.setText("Planning Week: Not Set")
        self.planning_week_widget.setStyleSheet("")
        self.setup_table()
        self.update_planned_weeks()
        self.update_all_combination_titles()
        self.update_ui_state()

    def update_remove_button(self):
        self.remove_button.setEnabled(len(self.file_list.selectedItems()) > 0)

    def setup_table(self):
        planning_horizons = ["W-1", "W-2", "W-3", "W-4", "W-5", "W-6", "W-7", "W-8", "W-9", "W-10"]
        self.planned_weeks_table.setRowCount(len(planning_horizons))
        self.planned_weeks_table.setColumnCount(4)
        self.planned_weeks_table.setHorizontalHeaderLabels(["Planning Horizon", "Amazon Week", "Available Planned Weeks", "Source Control"])

        for i, horizon in enumerate(planning_horizons):
            # Planning Horizon column
            item = QTableWidgetItem(horizon)
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.planned_weeks_table.setItem(i, 0, item)
            
            # Amazon Week column
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.planned_weeks_table.setItem(i, 1, item)
            
            # Available Planned Weeks column
            item = QTableWidgetItem("")
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            item.setBackground(QColor("yellow"))
            self.planned_weeks_table.setItem(i, 2, item)
            
            # Source Control column
            button = QPushButton("Select Source")
            button.setEnabled(False)
            self.planned_weeks_table.setCellWidget(i, 3, button)

        # Set resize mode for each column
        header = self.planned_weeks_table.horizontalHeader()
        for i in range(3):  # First three columns
            header.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)  # Source Control column

        self.planned_weeks_table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)

        # Adjust table height to fit content
        self.adjust_table_height()

    def adjust_table_height(self):
        # Calculate the total height of all rows plus the header
        total_height = self.planned_weeks_table.horizontalHeader().height()
        for i in range(self.planned_weeks_table.rowCount()):
            total_height += self.planned_weeks_table.rowHeight(i)

        # Add a small buffer (e.g., 2 pixels) to account for borders
        total_height += 2

        # Set the fixed height of the table
        self.planned_weeks_table.setFixedHeight(total_height)

    def create_combination_widget(self, index, default_preset):
        group = QGroupBox(f"Combination {index + 1}")
        group.setSizePolicy(QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Preferred)
        layout = QVBoxLayout()

        # Dropdown for preset selections
        preset_dropdown = QComboBox()
        preset_dropdown.addItems(["A", "B", "C", "D", "All", "Custom"])
        preset_dropdown.setCurrentText(default_preset)
        layout.addWidget(preset_dropdown)

        # Week range selection
        range_layout = QHBoxLayout()
        start_week = QSpinBox()
        start_week.setRange(1, 10)
        end_week = QSpinBox()
        end_week.setRange(1, 10)
        range_layout.addWidget(QLabel("From:"))
        range_layout.addWidget(start_week)
        range_layout.addWidget(QLabel("To:"))
        range_layout.addWidget(end_week)
        layout.addLayout(range_layout)

        # Title label
        title_label = WrappingLabel()
        title_label.setAlignment(Qt.AlignmentFlag.AlignTop | Qt.AlignmentFlag.AlignLeft)
        layout.addWidget(title_label)

        # Generate checkbox
        generate_checkbox = QCheckBox("Generate")
        generate_checkbox.setChecked(True)  # Set to True by default
        layout.addWidget(generate_checkbox)

        group.setLayout(layout)

        # Connect signals
        preset_dropdown.currentIndexChanged.connect(lambda: self.update_combination_range(group, index))
        start_week.valueChanged.connect(lambda: self.update_combination_title(group, index))
        end_week.valueChanged.connect(lambda: self.update_combination_title(group, index))
        generate_checkbox.stateChanged.connect(lambda: self.update_combination_title(group, index))

        self.update_combination_range(group, index)

        return group
    
    def resizeEvent(self, event: QResizeEvent):
        super().resizeEvent(event)
        if hasattr(self, 'update_combination_widths'):
            self.update_combination_widths()

    def update_combination_range(self, group, index):
        preset_dropdown = group.layout().itemAt(0).widget()
        start_week = group.layout().itemAt(1).itemAt(1).widget()
        end_week = group.layout().itemAt(1).itemAt(3).widget()
        
        preset = preset_dropdown.currentText()
        ranges = {
            "A": (2, 2), "B": (3, 5), "C": (6, 7), "D": (8, 10),
            "All": (2, 10), "Custom": (1, 10)
        }

        start, end = ranges[preset]
        start_week.setValue(start)
        end_week.setValue(end)

        self.check_range_validity(group, index)
        self.update_combination_title(group, index)

    def update_combinations_validity(self):
        for index, combination in enumerate(self.combinations):
            self.check_range_validity(combination, index)

    def check_range_validity(self, group, index):
        start_week = group.layout().itemAt(1).itemAt(1).widget()
        end_week = group.layout().itemAt(1).itemAt(3).widget()
        generate_checkbox = group.layout().itemAt(3).widget()

        if start_week.value() > end_week.value():
            generate_checkbox.setEnabled(False)
            return

        # Check if all weeks in the range are available
        all_weeks_available = all(
            self.planned_weeks_table.item(w-1, 2).text().startswith("w-")
            for w in range(start_week.value(), end_week.value() + 1)
        )

        generate_checkbox.setEnabled(all_weeks_available)
        if all_weeks_available and not generate_checkbox.isEnabled():
            generate_checkbox.setChecked(True)  # Re-enable and check if it becomes valid

    def update_all_combination_titles(self):
        for index, combination in enumerate(self.combinations):
            self.update_combination_title(combination, index)

    def update_combination_title(self, group, index):
        start_week = group.layout().itemAt(1).itemAt(1).widget().value()
        end_week = group.layout().itemAt(1).itemAt(3).widget().value()
        title_label = group.layout().itemAt(2).widget()
        generate_checkbox = group.layout().itemAt(3).widget()

        self.check_range_validity(group, index)

        weeks = ".".join(str(w) for w in range(start_week, end_week + 1))
        if self.planning_week is not None:
            title = f"summary_file_plwk{self.planning_week}_w-{weeks}"
        else:
            title = f"summary_file_plwk[Not Set]_w-{weeks}"
        
        if not generate_checkbox.isChecked():
            title += " (Disabled)"
        title_label.setText(title)

    def update_ui_state(self):
        has_files = self.file_list.count() > 0
        has_error = "Multiple Planning Weeks Detected" in self.planning_week_label.text()
        
        self.remove_button.setEnabled(has_files)
        self.clear_all_button.setEnabled(has_files)
        self.browse_button.setEnabled(not has_error)
        self.planned_weeks_table.setEnabled(not has_error)
        self.generate_button.setEnabled(has_files and not has_error)
        
        for combination in self.combinations:
            combination.setEnabled(not has_error)
        
        if has_files and self.planning_week is not None and not has_error:
            self.update_table()

    def remove_selected_files(self):
        for item in self.file_list.selectedItems():
            file_path = item.text()
            row = self.file_list.row(item)
            self.file_list.takeItem(row)
            if file_path in self.file_data:
                del self.file_data[file_path]
        
        # Reset planning week and hide warning if no files left
        if self.file_list.count() == 0:
            self.planning_week = None
            self.planning_week_label.setText("Planning Week: Not Set")
            self.planning_week_widget.setStyleSheet("")
        else:
            # Check if the error state needs to be cleared
            first_file = self.file_list.item(0).text()
            first_planning_week = self.extract_planning_week(os.path.basename(first_file))
            if all(self.extract_planning_week(os.path.basename(self.file_list.item(i).text())) == first_planning_week 
                for i in range(self.file_list.count())):
                self.planning_week = first_planning_week
                self.planning_week_label.setText(f"Planning Week: {self.planning_week}")
                self.planning_week_widget.setStyleSheet("")

        self.update_planned_weeks()
        self.update_all_combination_titles()  # Add this line
        self.update_ui_state()

    def update_table(self):
        if self.planning_week is None:
            return
        
        for i in range(10):
            amazon_week = (self.planning_week + i + 1) % 52
            if amazon_week == 0:
                amazon_week = 52
            item = QTableWidgetItem(str(amazon_week))
            item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.planned_weeks_table.setItem(i, 1, item)
        
        self.update_planned_weeks()

    def browse_files(self):
        files, _ = QFileDialog.getOpenFileNames(self, "Select Excel Files", "", "Excel Files (*.xlsx *.xls)")
        for file in files:
            if not self.file_list.file_exists(file):
                self.file_list.addItem(file)
                self.process_file(file)
            else:
                QMessageBox.warning(self, "Duplicate File", f"The file '{os.path.basename(file)}' has already been added.")

    def process_file(self, file_path):
        file_name = os.path.basename(file_path)
        extracted_planning_week = self.extract_planning_week(file_name)
        if extracted_planning_week:
            if self.planning_week is None:
                self.planning_week = extracted_planning_week
                self.planning_week_label.setText(f"Planning Week: {self.planning_week}")
                self.update_all_combination_titles()  # Add this line
            elif self.planning_week != extracted_planning_week:
                self.planning_week_label.setText("Planning Week: Multiple Planning Weeks Detected - Remove Summary File From Incorrect Planning Week")
                self.planning_week_widget.setStyleSheet("background-color: red; color: white; padding: 5px;")
        
        planned_horizons = self.extract_planned_horizons(file_name)
        self.file_data[file_path] = planned_horizons
        self.update_planned_weeks()
        self.update_combinations_validity()
        self.update_ui_state()

    def extract_planning_week(self, file_name):
        match = re.search(r"summary_file_plwk(\d+)", file_name)
        return int(match.group(1)) if match else None

    def extract_planned_horizons(self, file_name):
        match = re.search(r"w-([\d\.]+)", file_name)
        return [int(horizon) for horizon in match.group(1).split('.')] if match else []

    def update_planned_weeks(self):
        for i in range(10):
            self.planned_weeks_table.item(i, 2).setText("")
            self.planned_weeks_table.item(i, 2).setBackground(QColor("yellow"))
            self.planned_weeks_table.cellWidget(i, 3).setEnabled(False)

        for file_path, horizons in self.file_data.items():
            for horizon in horizons:
                if 1 <= horizon <= 10:
                    row = horizon - 1
                    current_text = self.planned_weeks_table.item(row, 2).text()
                    if current_text:
                        item = QTableWidgetItem("Duplicate")
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        item.setBackground(QColor("red"))
                        self.planned_weeks_table.setItem(row, 2, item)
                        button = self.planned_weeks_table.cellWidget(row, 3)
                        button.setEnabled(True)
                        button.clicked.connect(lambda checked, r=row: self.select_source(r))
                    else:
                        item = QTableWidgetItem(f"w-{horizon}")
                        item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                        item.setBackground(QColor("green"))
                        self.planned_weeks_table.setItem(row, 2, item)

    def select_source(self, row):
        dialog = QDialog(self)
        dialog.setWindowTitle("Select Source File")
        layout = QVBoxLayout()

        horizon = int(self.planned_weeks_table.item(row, 0).text().split('-')[1])
        files_with_horizon = [file for file, horizons in self.file_data.items() if horizon in horizons]

        for file in files_with_horizon:
            radio = QRadioButton(os.path.basename(file))
            layout.addWidget(radio)

        button_box = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        layout.addWidget(button_box)

        dialog.setLayout(layout)

        if dialog.exec() == QDialog.DialogCode.Accepted:
            for i in range(layout.count() - 1):  # -1 to exclude button box
                radio = layout.itemAt(i).widget()
                if radio.isChecked():
                    selected_file = [file for file in files_with_horizon if os.path.basename(file) == radio.text()][0]
                    horizon = int(self.planned_weeks_table.item(row, 0).text().split('-')[1])
                    item = QTableWidgetItem(f"w-{horizon}")
                    item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    item.setBackground(QColor(173, 216, 230))  # Light blue
                    self.planned_weeks_table.setItem(row, 2, item)
                    self.planned_weeks_table.cellWidget(row, 3).setEnabled(False)
                    # Disconnect the button to prevent multiple connections
                    self.planned_weeks_table.cellWidget(row, 3).clicked.disconnect()
                    break

    def get_enabled_combinations(self):
        enabled_combinations = []
        for combination in self.combinations:
            generate_checkbox = combination.layout().itemAt(3).widget()
            if generate_checkbox.isChecked() and generate_checkbox.isEnabled():
                start_week = combination.layout().itemAt(1).itemAt(1).widget().value()
                end_week = combination.layout().itemAt(1).itemAt(3).widget().value()
                title_label = combination.layout().itemAt(2).widget()
                title = title_label.text()
                enabled_combinations.append({
                    'start_week': start_week, 
                    'end_week': end_week,
                    'title': title
                })
        self.log_message(f"Identified {len(enabled_combinations)} enabled combinations")
        return enabled_combinations

    def get_file_paths(self):
        return [self.file_list.item(i).text() for i in range(self.file_list.count())]

    def start_combination_process(self):
        enabled_combinations = self.get_enabled_combinations()
        if not enabled_combinations:
            QMessageBox.warning(self, "No Combinations", "Please enable at least one valid combination before generating files.")
            return

        file_paths = self.get_file_paths()
        self.worker = FileCombinerWorker(file_paths, enabled_combinations, self.planning_week)
        self.worker.progress_updated.connect(self.update_progress)
        self.worker.error_occurred.connect(self.show_error)
        self.worker.process_completed.connect(self.show_process_completed)
        self.worker.save_location_requested.connect(self.get_save_location)
        self.worker.start()

        self.progress_dialog = QProgressDialog("Generating Combined Files", "Cancel", 0, 100, self)
        self.progress_dialog.setWindowTitle("Generating Combined Files")
        self.progress_dialog.setWindowModality(Qt.WindowModality.WindowModal)
        self.progress_dialog.setAutoReset(False)
        self.progress_dialog.setAutoClose(False)
        self.progress_dialog.canceled.connect(self.cancel_process)
        self.progress_dialog.show()

    def get_save_location(self):
        default_dir = r"W:\Shared With Me\11. OTR\01_ShareFolder\output to bigpush"
        save_dir = QFileDialog.getExistingDirectory(self, "Select Save Directory", default_dir)
        if save_dir:
            self.worker.set_save_directory(save_dir)
        else:
            self.worker.error_occurred.emit("Save location selection cancelled")

    def cancel_process(self):
        if self.worker.isRunning():
            self.worker.terminate()
            self.worker.wait()
            QMessageBox.information(self, "Process Cancelled", "The combination process has been cancelled.")
        self.progress_dialog.close()

    def show_process_completed(self, combination_names, save_directory, row_counts):
        self.progress_dialog.close()
        message = f"Process completed successfully.\n\nCombinations created:\n"
        for name in combination_names:
            base_name = os.path.splitext(name)[0]  # Remove file extension
            row_count = row_counts.get(base_name, "Unknown")
            message += f"{name} with {row_count} rows\n"
        message += f"\nSaved to: {save_directory}"
        QMessageBox.information(self, "Process Completed", message)

    def log_message(self, message):
        print(f"Log: {message}")  # This will print to the console

    def log_save_timing(self, timing_info):
        self.log_message(timing_info)

    def update_progress(self, value, message):
        self.progress_dialog.setValue(value)
        self.progress_dialog.setLabelText(message)

    def process_finished(self):
        self.progress_dialog.close()
        QMessageBox.information(self, "Process Completed", "All combinations have been processed and saved.")

    def show_error(self, error_message):
        QMessageBox.critical(self, "Error", error_message)
        self.progress_dialog.close()

    def show_process_completed(self, combination_names, save_directory):
        self.progress_dialog.close()
        message = f"Process completed successfully.\n\nCombinations created:\n"
        message += "\n".join(combination_names)
        message += f"\n\nSaved to: {save_directory}"
        QMessageBox.information(self, "Process Completed", message)

    def show_combination_completed(self, file_path):
        QMessageBox.information(self, "Combination Completed", f"Combined file saved as:\n{file_path}")
        self.progress_dialog.close()

    def restart(self):
        self.file_list.clear()
        self.file_data.clear()
        self.planning_week = None
        self.planning_week_label.setText("Planning Week: Not Set")
        self.planning_week_widget.setStyleSheet("")
        self.setup_table()
        for combination in self.combinations:
            self.update_combination_range(combination, self.combinations.index(combination))
        self.update_ui_state()  # Call this to update the UI state after restarting